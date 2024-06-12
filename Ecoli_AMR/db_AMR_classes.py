#!/usr/bin/env python
# Script to make a dictionary containing all classes & subclasses of antibiotics in reference database
# Result added in new worksheet "AMR_classes" of the Excel-file with the database
import pandas as pd
import sys
from openpyxl import Workbook
from openpyxl import load_workbook


# Check if any command-line arguments have been provided
if len(sys.argv) < 2:
    print("Error: Need to provide the correct command-line arguments.")
    print("Usage: python scriptname.py [1] [2] [3] [4] ...")
    print("\t[1] = Full path to csv-file with reference database")
    sys.exit(1)

# Store command-line argument(s) (=ST) as an integer in a list
refgens_db= sys.argv[1]

# Open the excel file and go to worksheet "refgenes_AMR"
wb = load_workbook(refgens_db)
ws_AMR = wb["refgenes_AMR"]

# Go through column 7 with "Class" and column 8 with "Subclass"
# Make a dictionary with antibiotics in the database : class as keys & subclass as value 
AB_dict = {}
for row in range(2, ws_AMR.max_row + 1):
    ab_class = ws_AMR.cell(row=row, column=7).value
    ab_subclass = ws_AMR.cell(row=row, column=8).value
    # If the class is already in the dictionary and the subclass is not in the list, add the subclass
    if ab_class in AB_dict:
        if ab_subclass not in AB_dict[ab_class]:
            AB_dict[ab_class].append(ab_subclass)
    # If the class is not in the dictionary, add it with the subclass as the first item in the list
    else:
        AB_dict[ab_class] = [ab_subclass]

# Make a new worksheet with classs & subclasses of AMR genes
ws_class = wb.create_sheet("AMR_classes")

# Add the dictionary to the worksheet : class in column 1 and subclass in column 2
row = 1
for key, value in AB_dict.items():
    ws_class.cell(row=row, column=1).value = key
    for val in value:
        ws_class.cell(row=row, column=2).value = val
        row += 1

# Save & close the workbook
wb.save(refgens_db)
wb.close()

# Print message for the user
print("New worksheet with classes & subclasses created in the Excel file.")