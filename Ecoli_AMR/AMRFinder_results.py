#!/usr/bin/env python
# Script to write results of AMRFinderPlus to an Excel-file
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# STEP 1 : CHECK & STORE COMMAND-LINE ARGUMENTS
################################################################################################################
# Check if any command-line arguments have been provided
if len(sys.argv) < 3:
    print("Error: Need to provide the correct command-line arguments.")
    print("Usage: python scriptname.py [1] ...")
    print("\t[1] = Full path to directory with AMRFinderPlus results")
    print("\t[2] = Full path to Excel file with AMR tools summary")
    sys.exit(1)

# Store command-line arguments
input_dir = sys.argv[1]
excel_path = sys.argv[2] # "/media/sf_SF/BIT11_Traineeship/Ecoli_WGS/AMR_tools_summary.xlsx"

# If input_dir has trailing slash, remove it
if input_dir[-1] == "/":
    input_dir = input_dir[:-1]
# path to result files : /home/guest/BIT11_Traineeship/Ecoli_AMR/Ecoli_WGS/AMR_output/AMR_output_excel

# STEP 2 : SUMMARIZE AMRFINDERPLUS RESULTS BY DETERMINING THE ANTIBIOTIC RESISTANCE & WRITE METADATA WITH GENES
################################################################################################################
# Open the copy of the result file
ws_summary_name = "Comparative AMR"
ws_genes_name = "AMRFinderPlus_genes"
wb_summary = load_workbook(excel_path)
ws_summary = wb_summary[ws_summary_name]
ws_genes = wb_summary[ws_genes_name]

# Row number where the output should be written to (on summary sheet & on gene sheet)
output_row = 3
gene_output_row = 3

# Get the list of files in the input directory
for file in os.listdir(input_dir):
    file_path = input_dir + "/" + file
    # Open the file with AMRFinderPlus results for each genome
    wb = load_workbook(file_path)
    ws = wb.active
    # Retrieve the genome name by removing "_output.xlsx" from the filenmae
    genome = file.replace("_output.xlsx" ,"")

    # Initialize flags for trimethoprim and sulfonamide detection
    found_trimethoprim = False
    found_sulfonamide = False

    # Read each line of the file
    for row in range(2, ws.max_row + 1):
        coverage = ws.cell(row=row, column=16).value
        # Determine what AMR gene has been detected & which antibiotic from the study it might cause resistance for
        #if ws.cell(row=row,column=9).value == "AMR" and coverage >95%:
        if ws.cell(row=row,column=9).value == "AMR":
            AMR_gene = ws.cell(row=row,column=6).value
            ab_class = ws.cell(row=row,column=11).value    
            ab_subclass = ws.cell(row=row,column=12).value
            #print(f"{file} : Gene = {AMR_gene}, Class = {ab_class}, Subclass = {ab_subclass}")
            if "FLUOROQUINOLONE" in ab_class and "FLUOROQUINOLONE" in ab_subclass:
                print(file, AMR_gene, ab_class, ab_subclass)
"""
            ###AMINOGLYCOSIDE###
            if  "AMINOGLYCOSIDE" in ab_class and  "AMINOGLYCOSIDE" in ab_subclass:
                # col 6 (amikacine) & 72 (tobramycin)
                ws_summary.cell(row=output_row, column=6, value="R")
                ws_summary.cell(row=output_row, column=72, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="AMINOGLYCOSIDE")
                ws_genes.cell(row=gene_output_row, column=4, value="AMINOGLYCOSIDE")
                gene_output_row += 1  
                #print(f"{file}: amikacin & tobramycin resistance with gene {AMR_gene}")
            if "AMINOGLYCOSIDE" in ab_class and "AMIKACIN" in ab_subclass:
                # col 6 (amikacine)
                ws_summary.cell(row=output_row, column=6, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="AMINOGLYCOSIDE")
                ws_genes.cell(row=gene_output_row, column=4, value="AMIKACIN")
                gene_output_row += 1                 
                #print(f"{file}: amikacin resistance with gene {AMR_gene}")
            if "AMINOGLYCOSIDE" in ab_class and "TOBRAMICIN" in ab_subclass:
                # col 72 (tobramycin)
                ws_summary.cell(row=output_row, column=72, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="AMINOGLYCOSIDE")
                ws_genes.cell(row=gene_output_row, column=4, value="TOBRAMICIN") 
                gene_output_row += 1 
                #print(f"{file}: tobramycin resistance with gene {AMR_gene}")

            ###BETA-LACTAM###
            if "BETA-LACTAM" in ab_class and  "BETA-LACTAM" in ab_subclass:
                # col 12 (amoxicillin), col 18 (amoxiclav), col 24 (aztreonam), col 30 (cefepime), col 36 (ceftazidime), col 54 (meropenem), col 60 (piperacillin-tazobactam)
                BL_cols = [12, 18, 24, 30, 36, 54, 60]
                for col in BL_cols:
                    ws_summary.cell(row=output_row, column=col, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="BETA-LACTAM")
                ws_genes.cell(row=gene_output_row, column=4, value="BETA-LACTAM") 
                gene_output_row += 1 
                #print(f"{file}: amoxicillin, amoxiclav, aztreonam, cefepime, ceftazidime, meropenem, piperacillin-tazobactam resistance with gene {AMR_gene}")
            if "BETA-LACTAM" in ab_class and "AMOXICILLIN-CLAVULANIC ACID" in ab_subclass:
                # col 18 (amoxiclav)
                ws_summary.cell(row=output_row, column=18, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="BETA-LACTAM")
                ws_genes.cell(row=gene_output_row, column=4, value="AMOXICILLIN-CLAVULANIC ACID") 
                gene_output_row += 1 
            if "BETA-LACTAM" in ab_class and "CEPHALOSPORIN" in ab_subclass:
                # col 36 (ceftazidime)
                ws_summary.cell(row=output_row, column=36, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="BETA-LACTAM")
                ws_genes.cell(row=gene_output_row, column=4, value="CEPHALOSPORIN") 
                gene_output_row += 1 
                #print(f"{file} ceftazidime resistance with gene {AMR_gene}")
            if "BETA-LACTAM" in ab_class and "CEFTAZIDIME-AVIBACTAM" in ab_subclass:
                # col 36 (ceftazidime)
                ws_summary.cell(row=output_row, column=36, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="BETA-LACTAM")
                ws_genes.cell(row=gene_output_row, column=4, value="CEFTAZIDIME-AVIBACTAM") 
                gene_output_row += 1 
                #print(f"{file} ceftazidime resistance with gene {AMR_gene}")
            if "BETA-LACTAM" in ab_class and  "CARBAPENEM" in ab_subclass:
                # col 54 (meropenem)
                ws_summary.cell(row=output_row, column=54, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="BETA-LACTAM")
                ws_genes.cell(row=gene_output_row, column=4, value="CARBAPENEM") 
                gene_output_row += 1 
                #print(f"{file}:meropenem resistance with gene {AMR_gene}")
            if "BETA-LACTAM" in ab_class and  "MEROPENEM" in ab_subclass:
                # col 54 (meropenem)
                ws_summary.cell(row=output_row, column=54, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="BETA-LACTAM")
                ws_genes.cell(row=gene_output_row, column=4, value="MEROPENEM") 
                gene_output_row += 1 
                #print(f"{file}:meropenem resistance with gene {AMR_gene}")
            if "BETA-LACTAM" in  ab_class and "PIPERACILLIN-TAZOBACTAM" in ab_subclass:
                # col 60 (piperacillin-tazobactam)
                ws_summary.cell(row=output_row, column=54, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="BETA-LACTAM")
                ws_genes.cell(row=gene_output_row, column=4, value="PIPERACILLIN-TAZOBACTAM")
                gene_output_row += 1 
                #print(f"{file}: piperacillin-tazobactam resistance with gene {AMR_gene}")

            ###FLUOROQUINOLONE###
            if "FLUOROQUINOLONE" in ab_class and "FLUOROQUINOLONE" in ab_subclass:
                # col 42 (ciprofloxacin)
                ws_summary.cell(row=output_row, column=42, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="FLUOROQUINOLONE")
                ws_genes.cell(row=gene_output_row, column=4, value="FLUOROQUINOLONE")
                gene_output_row += 1 
                #print(f"{file}: ciprofloxacin resistance with gene {AMR_gene}")

            ###COLISTIN###
            if "COLISTIN" in  ab_class and "COLISTIN" in ab_subclass:
                # col 48 (colistin)
                ws_summary.cell(row=output_row, column=48, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="COLISTIN")
                ws_genes.cell(row=gene_output_row, column=4, value="COLISTIN")
                gene_output_row += 1 
                #print(f"{file}: colistin resistance with gene {AMR_gene}")

            ###TETRACYCLINE###
            if "TETRACYCLINE" in ab_class and  "TETRACYCLINE" in ab_subclass:
                # col 66 (tigecyclin)
                ws_summary.cell(row=output_row, column=66, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="TETRACYCLINE")
                ws_genes.cell(row=gene_output_row, column=4, value="TETRACYCLINE")
                gene_output_row += 1 
                #print(f"{file}: tigecyclin resistance with gene {AMR_gene}")
            if "TETRACYCLINE" in ab_class and "TIGECYCLINE" in ab_subclass:
                # col 66 (tigecyclin)
                ws_summary.cell(row=output_row, column=66, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="TETRACYCLINE")
                ws_genes.cell(row=gene_output_row, column=4, value="TIGECYCLINE")
                gene_output_row += 1 
                #print(f"{file}: tigecyclin resistance with gene {AMR_gene}")
            if "MULTIDRUG" in ab_class and "TETRACYCLINE" in ab_subclass:
                # col 66 (tigecyclin)
                ws_summary.cell(row=output_row, column=66, value="R")
                ws_genes.cell(row=gene_output_row, column=1, value=genome)
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="MULTIDRUG")
                ws_genes.cell(row=gene_output_row, column=4, value="TETRACYCLINE")
                gene_output_row += 1 
                #print(f"{file}: tigecyclin resistance with gene {AMR_gene}")

            ###TRIMETHOPRIM-SULFAMETHOXAZOLE###
            if "TRIMETHOPRIM-SULFAMETHOXAZOLE" in ab_class and "TRIMETHOPRIM-SULFAMETHOXAZOLE" in ab_subclass:
                # col 78 (Trimethroprim_sulfamethoxazole)
                ws_summary.cell(row=output_row, column=78, value="R")
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="TRIMETHOPRIM-SULFAMETHOXAZOLE")
                ws_genes.cell(row=gene_output_row, column=4, value="TRIMETHOPRIM-SULFAMETHOXAZOLE")
                gene_output_row += 1 
                #print(f"{file}: Trimethroprim_sulfamethoxazole resistance with gene {AMR_gene}")
            if "TRIMETHOPRIM" in ab_class and "TRIMETHOPRIM" in ab_subclass:
                # Make a flag to check if both trimethoprim and sulfonamide resistance is detected and store the row number for later
                found_trimethoprim = True
                tri_sulf_row = output_row
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="TRIMETHOPRIM")
                ws_genes.cell(row=gene_output_row, column=4, value="TRIMETHOPRIM")
                gene_output_row += 1 
                #print(f"{file}: Trimethroprim resistance with gene {AMR_gene}")
            if "SULFONAMIDE" in ab_class and "SULFONAMIDE" in ab_subclass:
                # Make a flag to check if both trimethoprim and sulfonamide resistance is detected and store the row number for later
                found_sulfonamide = True
                tri_sulf_row = output_row
                ws_genes.cell(row=gene_output_row, column=2, value=AMR_gene)
                ws_genes.cell(row=gene_output_row, column=3, value="SULFONAMIDE")
                ws_genes.cell(row=gene_output_row, column=4, value="SULFONAMIDE")
                gene_output_row += 1 
                #print(f"{file}: Sulfamethoxazole resistance with gene {AMR_gene}")

    # In one input file or genome: if both trimethoprim and sulfonamide AMR genes are detected, write R in the summary sheet for Trimethroprim_sulfamethoxazole
    if found_trimethoprim == True and found_sulfonamide== True:
        # col 78 (Trimethroprim_sulfamethoxazole)
        ws_summary.cell(row=tri_sulf_row, column=78, value="R")

    # Increase the row number for the next genome
    #print(output_row)
    output_row +=1 

# Read the summary sheet and add S if there is no value (=resistance detected) for the AMRFinderPlus tool
for row in range(2, ws_summary.max_row + 1):
    AB_columns = [6, 12, 18, 24, 30, 36, 42, 48, 54, 60, 66, 72, 78]
    for col in AB_columns:
        if ws_summary.cell(row=row, column=col).value is None:
            ws_summary.cell(row=row, column=col, value="S")

# Read the gene sheet and remove empty rows
for row in range(ws_genes.max_row, 2, -1):  # Start from the last row and move upwards
    if ws_genes.cell(row=row, column=1).value is None:
        ws_genes.delete_rows(row)

# Save & close the workbook with summary results of AMRFinderPlus
wb_summary.save(excel_path)
wb_summary.close()
"""