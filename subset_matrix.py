#!/usr/bin/env python
# Script to make a new excel worksheet with a subset FastANI matrix based on the ANI results of one ST from an Excel file
import pandas as pd
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import itertools

# Check if any command-line arguments have been provided
if len(sys.argv) < 2:
	print("Error: Need to provide the correct command-line arguments.")
	print("Usage: python scriptname.py [1] ...")
	print("\t[1] ... = Specify one or more strain types (ST) (e.g. 262 1076), all genomes of those STs will be selected from the Excel file)")
	sys.exit(1)

# Store command-line argument(s) (=ST) as an integer in a list
ST_input = []
for arg in sys.argv[1:]:
    ST_input.append(int(arg))

# Load the workbook
excel_file_path = "/home/guest/BIT11_Traineeship/Scripts_traineeship/FastANI_matrix_Pseudomonas_aeruginosa (another copy).xlsx"
wb = load_workbook(excel_file_path)

# Select the worksheet named "MLST" & "Pseudomonas_aeruginosa"
ws_MLST=wb["MLST"]
ws_ANI=wb["Pseudomonas_aeruginosa"]

# Make a dictionary with the STs as key and GCFs as values
ST_dict = {}
for row in range(2, ws_MLST.max_row + 1):
    ST = ws_MLST.cell(row=row, column=2).value
    genome = ws_MLST.cell(row=row, column=1).value
    if ST in ST_dict:
        ST_dict[ST].append(genome)
    else:
        ST_dict[ST] = [genome]

# List of genomes for the subset ANI matrix
#GCF = ["GCF_033392255.1","GCF_013255565.1","GCF_001632245.1","GCF_030444495.1","GCF_029961345.1",
       #"GCF_027359235.1","GCF_019857465.1","GCF_036232165.1","GCF_030121895.1","GCF_021266605.1"]
        
# Make a list of all GCF values in the dictionary of the STs that were given as input (these genomes will be used for the subset ANI matrix)
GCF = []
for key, value in ST_dict.items():
     if key in ST_input:
         GCF.extend(value)

# Make a new worksheet for the subset ANI matrix
ws_ANI_subset = wb.create_sheet("Subset_ANI_matrix")

# Make a new worksheet for the subset MLST results (that were used for the subset ANI matrix)
ws_MLST_subset = wb.create_sheet("Subset_MLST_results")
# Write column headers
ws_MLST_subset.cell(row=1, column=1, value="RefSeq ID")
ws_MLST_subset.cell(row=1, column=2, value="ST")
        
# Write RefSeq identifiers from the GCF list to the new worksheets (ws_ANI_subset & ws_MLST_subset) & write the STs to the ws_MLST_subset worksheet
GCF_index = 0
index = 2
for genome in GCF:
    # Write the RefSeq ID as column and row headers to the ws_ANI_subset worksheet
    ws_ANI_subset.cell(row=index, column=1, value=genome)
    ws_ANI_subset.cell(row=1, column=index, value=genome)
    # Write the RefSeq ID as row header to the ws_MLST_subset worksheet
    ws_MLST_subset.cell(row=index, column=1, value=genome)
    # Write the STs of the genomes in the GCF list to the ws_MLST_subset worksheet
    for row in range(1, ws_MLST.max_row + 1):
        genome_MLST = ws_MLST.cell(row=row, column=1).value
        if genome_MLST == genome:
            ST = ws_MLST.cell(row=row, column=2).value
            if genome_MLST in GCF:
                ws_MLST_subset.cell(row=index, column=2, value=ST)
                GCF_index += 1
                index += 1

# Make all possible combinations of the genomes in the list
for genome_pair in itertools.product(GCF, GCF):
    # Extract the column- & row index of the genomes in the ANI matrix
    for row in range(2, ws_ANI.max_row+1):
            if (ws_ANI.cell(row, 1).value) == genome_pair[0]:
                row_index = row
    for col in range(2, ws_ANI.max_row+1):
            if (ws_ANI.cell(1, col).value) == genome_pair[1]:
                col_index = col           
    # Find the ANI-value of the genome pair in the Excel file
    ANI_value = ws_ANI.cell(row_index, col_index).value
    #print(genome_pair[0], genome_pair[1], ANI_value)
    # Extract the column - & row index of the genomes in the subset ANI matrix
    for row in range(2, ws_ANI_subset.max_row+1):
        if ws_ANI_subset.cell(row, 1).value == genome_pair[0]:
            row_index_subset = row
    for col in range(2, ws_ANI_subset.max_row+1):
        if ws_ANI_subset.cell(1, col).value == genome_pair[1]:
            col_index_subset = col
    # Write the ANI-value to the new worksheet
    ws_ANI_subset.cell(row_index_subset, col_index_subset, value=ANI_value)

# Save & close the workbook
wb.save(excel_file_path)
wb.close()
