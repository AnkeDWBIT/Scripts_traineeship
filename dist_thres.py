from shiny import App, render, ui, reactive
import pandas as pd
from sklearn.cluster import AgglomerativeClustering
from openpyxl import load_workbook
from collections import OrderedDict
from sklearn.metrics import adjusted_rand_score
import seaborn as sns 
import matplotlib.pyplot as plt 

# Shiny for python script that asks user for input of an excel file and select the worksheets with ANI & MLST results
# User can specify a value for distance_threshold via the slider
# Single-linkage agglomeratie clustering is performed
# Adjusted rand is computed to test the congruence between the clustering and the MLST results

# DEFINE THE UI OF THE APP
################################################################################################################
app_ui = ui.page_sidebar(
    # Sidebar for user input (upload Excel file & select worksheets for input data)
    ui.sidebar(
        ui.h2("Input file"),
        ui.input_file("file", "Upload Excel File", accept=".xlsx"),
        ui.h4("Select worksheets"),
        ui.input_select("select_ANI", "ANI matrix", choices=[],),
        ui.input_select("select_MLST", "MLST (metadata)", choices=[]),
    ),
    # Main content area
    ui.card(
        ui.card(
            ui.h2("Optimizing distance_threshold in Agglomerative Clustering of ANI values"),
            ui.input_slider("n", "Distance_threshold", min=0, max=1, value=0.3),
            ui.input_action_button("go", "Perform Clustering", class_="btn-success"),
            ),
        ui.card(
            ui.card_header("Congruence of partitioning methods (MLST & ANI clustering)"),
            ui.output_text("txt"),
        ),
        ui.layout_columns(
            # Output area where the user can see the results of the clustering
            ui.card(
                ui.card_header("Heatmap of ANI-values"),
                ui.output_plot("heatmap"),
                ),
            ui.card(
                ui.card_header("Results table: Ordered MLST results with ANI cluster labels"),
                ui.output_data_frame("ordered_MLST_table"),
            )
        )
    )
)

# DEFINE THE SERVER LOGIC OF THE APP
################################################################################################################
def server(input, output, session):
    session.df = None
    session.df_MLST = None
    session.labels = None
    data_processed = reactive.Value(False)

    @reactive.Effect
    @reactive.event(input.file) # Event that triggers the function when the file is uploaded
    # Function that opens the Excel file and finds the worksheets in it, used for the select input
    def update_sheet_choices():
        file_info = input.file()
        if file_info:
            # Extract the actual file path
            file_data = file_info[0]['datapath']
            # Read the Excel file and get the sheet names
            xls = pd.ExcelFile(file_data)
            worksheets = xls.sheet_names
            # Update the choices of the select input for sheets
            ui.update_select("select_ANI", choices=worksheets)
            ui.update_select("select_MLST", choices=worksheets)

    # Process data from the excel file (agglomerative clustering, combine results with metadata)
    @reactive.Effect
    @reactive.event(input.go) # Event that triggers the function when the button is clicked
    # Function that performs agglomerative clustering on the ANI values
    def process_data():
        file_info = input.file()
        ANI_sheet_name = input.select_ANI()
        MLST_sheet_name = input.select_MLST()
        
        if file_info and ANI_sheet_name and MLST_sheet_name:
            # Load the percent identity matrix from an Excel file into a dataframe
            df = pd.read_excel(file_info[0]['datapath'], sheet_name=ANI_sheet_name, engine='openpyxl', header=0, index_col=0)
            # Convert the DataFrame to a Numpy array & Convert percent identity to distance (assuming values are in percentage form)
            percent_identity_matrix = df.to_numpy()
            distance_matrix = 100 - percent_identity_matrix

            # Perform agglomerative clustering
            dist_thres = input.n()
            clustering = AgglomerativeClustering(metric='precomputed', linkage='average', distance_threshold=dist_thres, n_clusters=None)
            clustering.fit(distance_matrix)
            
            # Get the cluster labels
            labels = clustering.labels_

             # Load the data from the Excel file
            wb = load_workbook(file_info[0]['datapath'])
            ws_MLST=wb[MLST_sheet_name]
            ws_ANI=wb[ANI_sheet_name]
            
            # Store the genomes of the ANI sheet in a list
            ANI_genomes = []
            for row in range(2, ws_ANI.max_row + 1):
                ANI_genomes.append(ws_ANI.cell(row=row, column=1).value)
            
            # Store the genomes and STs of the MLST sheet in a dictionary
            MLST_dict = {}
            for row in range(2, ws_MLST.max_row + 1):
                MLST_dict[ws_MLST.cell(row=row, column=1).value] = ws_MLST.cell(row=row, column=2).value

            # Create an ordered dictionary (MLST dictionary in order of ANI genomes list)
            ordered_MLST = OrderedDict((key, MLST_dict[key]) for key in ANI_genomes)
            # Convert ordered dictionary to a DataFrame for display
            df_MLST = pd.DataFrame(list(ordered_MLST.items()), columns=['Genome', 'ST'])
            # Add the cluster labels to the DataFrame
            df_MLST['Cluster'] = labels

            # Store these results in the session for use in other functions
            session.df_MLST = df_MLST
            session.df = df
            session.labels = labels

             # Set data_processed to True
            data_processed.set(True)
    
    # Render the output of the app after the data is processed
    
    # Make an output table with the MLST results and ANI cluster labels, ordered as the genomes in the ANI matrix
    @render.data_frame
    def ordered_MLST_table():
        #print(session.df_MLST)
        if data_processed.get() and session.df_MLST is not None:
            #print("Rendering ordered_MLST_table")  # Log for debugging
            return session.df_MLST
        return pd.DataFrame()

    # Add a text output with the adjusted rand value to the output area
    @render.text
    def txt():  
        if data_processed.get() and session.df_MLST is not None and session.labels is not None:
            adjusted_rand = adjusted_rand_score(session.df_MLST['ST'], session.labels) # Calculate the adjusted rand value (congruence between two typin methods)
            return (f"Agglomerative clustering of ANI values with distance_threshold = {input.n()}"
                    f"Data retrieved from the Excel file on worksheets '{input.select_ANI()}' & '{input.select_MLST()}'"
                    f"Clustering results summarized in the table"
                    f"Clustering congruence with MLST: {round(adjusted_rand, 4)}")
        return "No data to display yet."
      
    @render.plot
    def heatmap():
        if data_processed.get() and session.df is not None:
            #print("Rendering heatmap")  # Log for debugging
            # Load the ANI data in a dataframe
            ANI_df = session.df
            # Make a heatmap from the dataframe
            plot = sns.heatmap(ANI_df, cmap="vlag", xticklabels=True, yticklabels=True, cbar_kws={'label': 'Average Nucleotide Identity (%)'})
            """
            # Add bars with metadata to the figure (e.g. MLST or ANI-clustering per genome)
            if session.df_MLST is not None:
                for index, genome in enumerate(ANI_df.index, start=0):
                    ST = session.df_MLST.loc[session.df_MLST['Genome'] == genome, 'ST'].values[0]
                    plt.gca().add_patch(plt.Rectangle((1, index), 1, 1, color='C'+str(ST)))
            """
            return plot.figure
        return None

# Create the Shiny app
app = App(app_ui, server)