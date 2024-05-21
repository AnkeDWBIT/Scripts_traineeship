#!/usr/bin/env python
# Script for quality control of ANI matrix
# Look for genomes that consistenly have a percent identity lower than 95 with all other genomes of the dataset (indicative of a diferent species)
# Make new worksheet for ANI and MLST results where genomes with doubtful species identity are removed

import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import itertools

# PROCESS INPUT ARGUMENTS
######################################################################################################################################
# Check if any command-line arguments have been provided
if len(sys.argv) < 3:
    print("Error: Need to provide the correct command-line arguments.")
    print("Usage: python scriptname.py [1] [2] [3]")
    print("\t[1] = Full path to Excel file with ANI matrix and MLST results")
    print("\t[2] = Worksheet with ANI matrix in Excel file [1] e.g. Pseudomonas_aeruginosa")
    print("\t[3] = Worksheet with MLST results in Excel file [1] e.g. MLST")
    sys.exit(1)

# Store command-line argumentS
excel_file_path = sys.argv[1]
ANI_sheet_name = sys.argv[2]
MLST_sheet_name = sys.argv[3]

# LOAD THE ANI MATRIX FROM THE EXCEL FILE  
######################################################################################################################################
# Load the workbook
#excel_file_path = "/home/guest/BIT11_Traineeship/Scripts_traineeship/FastANI_matrix_Pseudomonas_aeruginosa_4thcopy.xlsx"
wb = load_workbook(excel_file_path)
# Select the worksheet named "MLST" & "Pseudomonas_aeruginosa"
ws_MLST=wb[MLST_sheet_name]
ws_ANI=wb[ANI_sheet_name]

# LOOK FOR GENOMES WITH AN AVERAGE ANI VALUE BELOW 95
######################################################################################################################################
# Iterate over each row in the matrix except row 1 (header row) and the first values in each row (header values)
genomes_to_remove = {}
for row in range(2, ws_ANI.max_row + 1):
    # Look for rows where the average ANI value is below 0.95
    genome_name = ws_ANI.cell(row=row, column=1).value
    ani_values = [cell.value for cell in ws_ANI[row][2:]]  # Exclude the first cell
    avg_ANI = sum(ani_values) / len(ani_values)
    if avg_ANI < 95:
        # Add the genome name and average ANI value to a dictionary (key = genome name, value = average ANI value) if avg_ANI < 95
        genomes_to_remove[genome_name] = avg_ANI

# MAKE A NEW FILE WITH QC RESULTS OF THE GENOMES
######################################################################################################################################
# If the dictionary is empty, add a short statement to the file
# If the dictionary is not empty, write the genome names and average ANI values to the file
if not genomes_to_remove:
    with open("genomes_to_remove.txt", "w") as file:
         file.write("No genomes were found with an average ANI value below 95.")
else:
        with open("genomes_to_remove.txt", "w") as file:
            file.write("Genomes to remove \n(average ANI <95 is indicative of a different species than other genomes in the dataset):\n\n")
            for genome, avg_ANI in genomes_to_remove.items():
                 file.write(f"{genome}\t{avg_ANI}\n")


# MAKE A NEW WORKSHEETS FOR ANI AND MLST RESULTS WHERE GENOMES WITH DOUBTFUL SPECIES IDENTITY ARE REMOVED
######################################################################################################################################
# Make a new worksheet for the reviewed ANI matrix
ws_ANI_reviewed = wb.create_sheet("ANI_matrix_reviewed")
# Copy the original ANI matrix to the new worksheet
for row in ws_ANI.iter_rows(values_only=True):
    ws_ANI_reviewed.append(row)
    # If the genome name is a key in the dictionary, remove the row
    for key, value in genomes_to_remove.items():
         if key == row[0]:
             ws_ANI_reviewed.delete_rows(ws_ANI_reviewed.max_row)
             # If the genome name is in the dictionary, remove the column
             for cell in row[1:]:
                  if cell == row[0]:
                      ws_ANI_reviewed.delete_cols(cell)

# Make a new worksheet for the reviewed MLST results
ws_MLST_reviewed = wb.create_sheet("MLST_reviewed")
# Copy the original MLST results to the new worksheet
for row in ws_MLST.iter_rows(values_only=True):
    ws_MLST_reviewed.append(row)
    # If the genome name is a key in the dictionary, remove the row
    for key, value in genomes_to_remove.items():
         if key == row[0]:
             ws_MLST_reviewed.delete_rows(ws_MLST_reviewed.max_row)

# Save & close the workbook
wb.save(excel_file_path)
wb.close()