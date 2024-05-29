#!/usr/bin/env python
# Script to make new ANI & MLST excel worksheets ("Subset_ANI_matrix_3STs" & "Subset_MLST_3STs") 
# The subset consists of genomes that have an ST that occurs at least 3 times in the dataset
from openpyxl import load_workbook
import sys
import itertools

# STEP 1 : HANDLE COMMAND-LINE ARGUMENTS
############################################################################################################
# Check if any command-line arguments have been provided
if len(sys.argv) < 4:
    print("Error: Need to provide the correct command-line arguments.")
    print("Usage: python scriptname.py [1] [2] [3]")
    print("\t[1] = Full path to Excel file with ANI matrix and MLST results")
    print("\t[2] = Worksheet with ANI matrix in Excel file [1] e.g. Pseudomonas_aeruginosa or ANI_matrix_reviewed")
    print("\t[3] = Worksheet with MLST results in Excel file [1] e.g. MLST or MLST_reviewed")
    sys.exit(1)

# Store command-line arguments
excel_file_path = sys.argv[1]
ANI_sheet_name = sys.argv[2]
MLST_sheet_name = sys.argv[3]

# STEP 2 : PROCESS EXCEL FILE
############################################################################################################
# Load the workbook & some of it's worksheets
#excel_file_path = "/home/guest/BIT11_Traineeship/Scripts_traineeship/FastANI_matrix_Pseudomonas_aeruginosa_results.xlsx"
wb = load_workbook(excel_file_path)
ws_MLST=wb[MLST_sheet_name]
ws_ANI=wb[ANI_sheet_name]

# Make a dictionary with the STs as key and genomes as values
ST_dict = {}
for row in range(2, ws_MLST.max_row + 1):
    ST = ws_MLST.cell(row=row, column=2).value
    genome = ws_MLST.cell(row=row, column=1).value
    if ST in ST_dict:
        ST_dict[ST].append(genome)
    else:
        ST_dict[ST] = [genome]

# Reduce the dictionary to only the genomes with STs that occur at least 3 times in the dataset
ST_dict = {key: value for key, value in ST_dict.items() if len(value) >= 3} 

# Make a list of all genomes (values) in the dictionary
genomes = []
for value in ST_dict.values():
    genomes.extend(value)

# STEP 3 : MAKE NEW WORKSHEETS IN THE EXCEL FILE WITH REDUCED DATA -> ONLY GENOMES WITH STs THAT OCCUR AT LEAST 3 TIMES
############################################################################################################
# Make a new worksheet for the subset ANI matrix & MLST results
ws_ANI_subset = wb.create_sheet("Subset_ANI_matrix_3STs")
ws_MLST_subset = wb.create_sheet("Subset_MLST_3STs")
# Write column headers to the new MLST worksheet
ws_MLST_subset.append(["RefSeq ID", "ST"])
"""
ws_MLST_subset.cell(row=1, column=1, value="RefSeq ID")
ws_MLST_subset.cell(row=1, column=2, value="ST")
"""
# Write RefSeq identifiers from the genomes list to the new worksheets & write the STs to the ws_MLST_subset worksheet
GCF_index = 0
index = 2
for genome in genomes:
    # Write the RefSeq ID as column and row headers to the ws_ANI_subset worksheet
    ws_ANI_subset.cell(row=index, column=1, value=genome)
    ws_ANI_subset.cell(row=1, column=index, value=genome)
    # Write the RefSeq ID as row header to the ws_MLST_subset worksheet
    ws_MLST_subset.cell(row=index, column=1, value=genome)
    # Write the STs of the genomes in the GCF list to the ws_MLST_subset worksheet
    for row in range(2, ws_MLST.max_row + 1):
        genome_MLST = ws_MLST.cell(row=row, column=1).value
        if genome_MLST == genome:
            ST = ws_MLST.cell(row=row, column=2).value
            if genome_MLST in genomes:
                ws_MLST_subset.cell(row=index, column=2, value=ST)
                GCF_index += 1
                index += 1

# Pre-compute the indices of the genomes in the original ANI matrix
genome_row_indices = {ws_ANI.cell(row, 1).value: row for row in range(2, ws_ANI.max_row + 1)}
genome_col_indices = {ws_ANI.cell(1, col).value: col for col in range(2, ws_ANI.max_column + 1)}

# Pre-compute the indices of the genomes in the subset ANI matrix
subset_genome_row_indices = {ws_ANI_subset.cell(row, 1).value: row for row in range(2, ws_ANI_subset.max_row + 1)}
subset_genome_col_indices = {ws_ANI_subset.cell(1, col).value: col for col in range(2, ws_ANI_subset.max_column + 1)}

# Make all possible combinations of the genomes in the list
for genome_pair in itertools.product(genomes, genomes):
    # Extract the ANI value
    row_index = genome_row_indices[genome_pair[0]]
    col_index = genome_col_indices[genome_pair[1]]
    ANI_value = ws_ANI.cell(row_index, col_index).value
    # Write the ANI-value to the new worksheet
    row_index_subset = subset_genome_row_indices[genome_pair[0]]
    col_index_subset = subset_genome_col_indices[genome_pair[1]]
    ws_ANI_subset.cell(row_index_subset, col_index_subset, value=ANI_value)

# STEP 4 : MAKE NEW WORKSHEETS IN THE EXCEL FILE WITH REDUCED DATA -> REMOVE GENOMES WITH UNKNOWN STs
############################################################################################################
# Copy the worksheets with the subset data
ws_ANI_copy = wb.copy_worksheet(ws_ANI_subset)
ws_MLST_copy = wb.copy_worksheet(ws_MLST_subset)
# Rename the new worksheets
ws_ANI_copy.title = "ANI_excl_unknown"
ws_MLST_copy.title = "MLST_excl_unknown"
# Store the row indeces of the genomes with Unknown STs
unknown_STs = []
for row in range(2, ws_MLST_copy.max_row + 1):
    ST = ws_MLST_copy.cell(row=row, column=2).value
    if ST == "Unknown":
        unknown_STs.append(row)
# Remove the rows (&columns) with Unknown STs from the ANI & MLST worksheets
for row in unknown_STs:
    ws_ANI_copy.delete_rows(row)
    ws_ANI_copy.delete_cols(row)
    ws_MLST_copy.delete_rows(row)
    
# Save & close the workbook
wb.save(excel_file_path)
wb.close()
