#!/usr/bin/env python
# Script to extract 16S rRNA sequences from .gbff files of genomes listed in the file "genomes_to_remove.txt" and write them to a file "QC_rRNAs.fa",
# then installs the 16S_ribosomal_RNA databse in "05_BLAST_16S_rRNA_database" (if not done yet) & runs BLAST on the sequences from "QC_rRNAs.fa" against the database,
# results are saved in "05_BLAST/blast_results" directory, one file per sequence
# Dependency : BLAST+ needs to be installed on the system (e.g. in a conda environment that is active when executing the script)
import sys
import os
from Bio import SeqIO
from Bio.SeqRecord import SeqRecord     
import subprocess
import openpyxl
from openpyxl.utils import get_column_letter

# STEP 1 : HANDLE COMMAND-LINE ARGUMENTS
############################################################################################################
# Check if any command-line arguments have been provided
if len(sys.argv) < 4:
    print("Error: Need to provide a command-line argument.")
    print("Usage: python scriptname.py [1] [2] [3]")
    print("\t[1] = Full path to directory with .gbff files to extract 16S rRNA sequences from (e.g. /home/guest/BIT11_Traineeship/01_Paeruginosa_refseq_genomes/ncbi_dataset/data/)")
    print("\t[2] = Full path to genomes_to_remove.txt file (holds genomes with avg ANI <95%)")
    print("\t[3] = Full path where to make BLAST output directory, not including the BLAST output dir (e.g. /home/guest/BIT11_Traieeship/)")
    sys.exit(1)

# Store the command-line argument(s) in an object
gbff_dir = sys.argv[1]
genomes_file = sys.argv[2]
blast_dir = sys.argv[3]

#gbff_dir = "/home/guest/BIT11_Traineeship/01_Paeruginosa_refseq_genomes/ncbi_dataset/data/"
#genomes_file = "/home/guest/BIT11_Traineeship/Scripts_traineeship/genomes_to_remove.txt"
#blast_dir = "/home/guest/BIT11_Traineeship/"

# STEP 2 : Extract genomes from the "genomes_to_remove.txt" file & construct filenames of .gbff files to look for
############################################################################################################
gbff_files = []
genomes = []
# Open the file with genomes to remove (location specified as input argument [2])
with open(genomes_file, 'r') as file:
	for line in file:
		# Look for lines starting with "GCF_" (RefSeq identifiers)
		if line.startswith("GCF_"):
			genome = line.split("\t")[0] # Extract the RefSeq ID out of the line
			gbff_file = genome + ".gbff"
			gbff_files.append(gbff_file)
			genomes.append(genome)
#print(gbff_files)
        
# STEP 3 : MAKE OUTPUT DIRECTORIES
############################################################################################################
# Make a new directory "05_BLAST" with subdirectories "blast_results" & "16S_rRNA_database" (location specified as input argument [3])
# If the database is already installed, don't install it again
path_db = os.path.join(blast_dir, "05_BLAST/16S_rRNA_database/")
if path_db in os.listdir(blast_dir):
	print("16S ribosomal RNA database already installed.")
else:
	os.makedirs(path_db, exist_ok=True)
path_results = os.path.join(blast_dir, "05_BLAST/blast_results/")
os.makedirs(path_results, exist_ok=True)

# STEP 4 : CREATE A FILE "QC_rRNAs.fa" WITH 16Ss rRNA SEQUENCES OF GENOMES FROM "genomes_to_remove.txt"
###########################################################################################################
os.chdir(path_results)
output_handle=open("QC_rRNAs.fa","w")
sixteen_s_records=[]
seq_nb = 1 # Add a sequence number to the 16S rRNA sequence name to avoid duplicate names (when one genome has multiple 16S rRNA sequences)
# Look in the specified input directory [1] for the .gbff files (paths derived from the genomes_to_remove.txt file)
for root, dirs, files in os.walk(gbff_dir):
	for file in files:
		if file in gbff_files:
			name = os.path.splitext(os.path.basename(file))[0]
			gbff_file_path = os.path.join(gbff_dir,name,file)
			with open(gbff_file_path, 'r') as gbff:
				for record in SeqIO.parse(gbff, 'genbank'):
					for feature in record.features:
						if(feature.type == "rRNA"):
							if '16S ribosomal RNA' in feature.qualifiers['product'][0]:
								sequence = feature.extract(record.seq)
								sixteen_s_records.append(SeqRecord(sequence, id=f"{name}_16S_{record.id}_seq{seq_nb}", description=""))
								seq_nb += 1
# Write the sequences to the output file
SeqIO.write(sixteen_s_records, output_handle, "fasta")
output_handle.close()

# STEP 5 : INSTALLATION OF 16S-ribosomal-RNA DATABASE FOR NCBI BLAST
###########################################################################################################
# Move to the "16S_rRNA_database" directory, download the 16S_ribosomal_RNA database from NCBI and decompress it
database = "16S_ribosomal_RNA"
os.chdir(path_db)
subprocess.run(["update_blastdb.pl", "--decompress", database])
# Check if the database has been installed correctly
results = subprocess.run(["blastdbcheck", "-db", database], capture_output=True, text=True)
# Write the output of the database installation to a file
with open("blastdbcheck_output.txt", "w") as out_file:
	out_file.write(results.stdout)
	out_file.write(results.stderr)
# If the output file doesn't have errors, print a message to the user
if not results.stderr:
	print("16S ribosomal RNA database installed successfully.")
	print(results.stdout)
else:
	print("Error: 16S ribosomal RNA database installation failed.")
	print(results.stderr)

# STEP 6 : RUN NCBI BLAST WITH THE SEQUENCES FROM "QC_rRNAs.fa"
###########################################################################################################
# In the output folder, make an excel file for each genome
for genome in genomes:
	wb = openpyxl.Workbook()
	wb.save(os.path.join(path_results,f"blast_{genome}.xlsx"))
	wb.template = True
# Open the file with the rRNA sequences and iterate over each sequence
with open(os.path.join(path_results,"QC_rRNAs.fa"), "r") as file:
	for record in SeqIO.parse(file, "fasta"):
		print(f"Running BLAST for sequence ID: {record.id}")	
        # Write the sequence to a temporary file
		temp_input = os.path.join(path_results,f"temp_{record.id}.fasta")
		with open(temp_input, "w") as temp_file:
			temp_file.write(f">{record.id}\n{record.seq}\n")
        # Construct the blastn command with custom output format
		blastn_cmd = [
            "blastn",
            "-query", temp_input,
            "-db", database,
            "-outfmt", "6 qseqid sseqid pident staxids sacc sscinames stitle"
        ]
        # Run blastn & capture the results
		results = subprocess.run(blastn_cmd, capture_output=True, text=True)
		# Open the excel file with the genome name
		for genome in genomes:
			if genome in record.id:
				output_file = os.path.join(path_results,f"blast_{genome}.xlsx")
				# Make a new worksheet for the record.id
				wb = openpyxl.load_workbook(output_file)
				ws = wb.create_sheet(record.id)
				# Write the header and BLAST results (output or error) to the output file
				header = "query_id\tsubject_id\tpercent_identity\ttaxonomy_id\taccession\tscientific_name\tsubject_title\n"
				ws.append(header.split("\t"))
				if results.stderr:
					ws.append(results.stderr.split("\n"))
					print(f"Error: {results.stderr}")
				else:
					for line in results.stdout.split("\n"):
						ws.append(line.split("\t"))
				# Autofit the column width
				for column_cells in ws.columns:
					length = max(len(str(cell.value)) for cell in column_cells)
					ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
				# Remove empty sheet with name "Sheet"
				if "Sheet" in wb.sheetnames:
					del wb["Sheet"]
				# Save & close the workbook
				wb.save(output_file)
				# Remove the temporary input file
				subprocess.run(["rm", temp_input])