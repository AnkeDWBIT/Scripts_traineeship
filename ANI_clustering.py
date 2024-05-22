#!/usr/bin/env python
# Script to perform Single-Linkage Agglomerative Clustering of ANI (distance) values with different values for distance_threshold
# New worksheet (MLST_ordered) where MLST results are ordered to match the order of the genomes in the ANI matrix
# Clustering results are written to this new worksheet (one column per tested distance_threshold value)
import pandas as pd
import sys
import pandas as pd
from sklearn.cluster import AgglomerativeClustering
from openpyxl import load_workbook

# PROCESS INPUT ARGUMENTS
######################################################################################################################################
# Check if any command-line arguments have been provided
if len(sys.argv) < 4:
    print("Error: Need to provide the correct command-line arguments.")
    print("Usage: python scriptname.py [1] [2] [3] [4] ...")
    print("\t[1] = Full path to Excel file with ANI matrix and MLST results")
    print("\t[2] = Worksheet with ANI matrix in Excel file [1] e.g. Pseudomonas_aeruginosa or ANI_matrix_reviewed")
    print("\t[3] = Worksheet with MLST results in Excel file [1] e.g. MLST or MLST_reviewed")
    print("\t[4] ... = Values to test for the distance threshold (e.g. 0.5 0.05)")
    sys.exit(1)

# Store command-line arguments
excel_file_path = sys.argv[1]
ANI_sheet_name = sys.argv[2]
MLST_sheet_name = sys.argv[3]
values_range = sys.argv[4:]

# Store the values_range as floats
values_range = [float(i) for i in values_range]

# LOAD THE ANI & MLST DATA FROM THE EXCEL FILE  
######################################################################################################################################
#excel_file_path = "/home/guest/BIT11_Traineeship/Scripts_traineeship/FastANI_matrix_Pseudomonas_aeruginosa_copy.xlsx"
#ANI_sheet_name = "Pseudomonas_aeruginosa"
#MLST_sheet_name = "MLST"
# Load the workbook
wb = load_workbook(excel_file_path)
# Select the worksheet with MLST and ANI results
ws_MLST=wb[MLST_sheet_name]
ws_ANI=wb[ANI_sheet_name]

# MAKE A NEW WORKSHEET WHERE ORDER OF THE MLST RESULTS MATCH THE GENOME ORDER IN THE ANI MATRIX
######################################################################################################################################
# Make a new worksheet for the ordered MLST results (& later for the clustered ANI values)
ws_ord_clus = wb.create_sheet("Ordered_Clustered")
# Write column headers
ws_ord_clus.cell(row=1, column=1, value="RefSeq ID")
ws_ord_clus.cell(row=1, column=2, value="ST")

# In a new worksheet, paste the genomes as orderd in ws_ANI and the corresponding MLST type from ws_MLST
for ANI_row in range(2, ws_ANI.max_row + 1):
    ANI_genome = ws_ANI.cell(row=ANI_row, column=1).value
    # Link MLST_genome to MLST_row
    for MLST_row in range(2, ws_MLST.max_row + 1):
        MLST_genome = ws_MLST.cell(row=MLST_row, column=1).value
        # If ANI_genome and MLST_genome are the same, write the genome
        if ANI_genome == MLST_genome:
            ws_ord_clus.cell(row=ANI_row, column=1, value=ANI_genome)
            ws_ord_clus.cell(row=ANI_row, column=2, value=ws_MLST.cell(MLST_row,2).value)

         
# AGGLOMERATIVE CLUSTERING OF ANI VALUES
# WRITE RESULTS TO EXCEL AND COMPARE PER GENOME IF ANI CLUSTER & MLST TYPE CORRESPOND
######################################################################################################################################
# Load the percent identity matrix from an Excel file into a dataframe
percent_identity_df = pd.read_excel(excel_file_path, sheet_name=ANI_sheet_name, index_col=0)
# Convert the DataFrame to a Numpy array & Convert percent identity to distance (assuming values are in percentage form)
percent_identity_matrix = percent_identity_df.to_numpy()
distance_matrix = 100 - percent_identity_matrix

# Perform agglomerative clustering
# Find the best value for the distance threshold by testing a range of values
col = 3
#for value in range(5, 0, -1):
    #values_range = value/10
#values_range = [0.5, 0.05]
for value in values_range:
    clustering = AgglomerativeClustering(metric='precomputed', linkage='average', distance_threshold=value, n_clusters=None)
    clustering.fit(distance_matrix)
    # Get the cluster labels
    labels = clustering.labels_
    #print("ANI cluster labels (using distance_threshold = ", value, ")", labels)
    # Add the labels from the Agglomerative clustering of ANI values to the ws_ord_clus worksheet staring form row 2, column 3
    header = "ANI cluster (distance_threshold = " + str(value) + ")"
    ws_ord_clus.cell(row=1, column=col, value=header)
    index = 2
    for label in labels:
        ws_ord_clus.cell(row=index, column=col, value=label)
        index += 1
    col += 1

# Save & close the workbook
wb.save(excel_file_path)
wb.close()

# MESSAGE WHEN SCRIPT IS FINISHED
######################################################################################################################################
print("Script finished.\nANI clustering results have been added to the specified Excel file on worsheet 'Ordered_Clustered'.\nThe distance threshold values tested were: ", values_range)